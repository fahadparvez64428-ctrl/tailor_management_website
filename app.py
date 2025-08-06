from flask import Flask, render_template, request, redirect, url_for
from flask import Flask, render_template, request, redirect, url_for, flash, session

import os
import openpyxl
from datetime import datetime
import pandas as pd
from flask import send_file
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
app = Flask(__name__)
app.secret_key = "tailor-secret"  # Add this line



# ========== Dummy Login Credentials ==========
USERNAME = "admin"
PASSWORD = "1234"

# ========== Excel File Path Setup ==========
# ========== Excel File Path Setup ==========

EXCEL_FILE_PATH = os.path.join(os.path.dirname(__file__), "customers.xlsx")


# Ensure file exists before loading
if not os.path.exists(EXCEL_FILE_PATH):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "ID", "Name", "Phone", "Address",
        "collar", "ban", "gara_Goal", "gara_churas",
        "fornt_poket", "site_poket", "single_salay", "duble_salay", "pajama",
        "shalwar_checkbox", "Option11", "Option12",
        "kameez_Length", "bazu", "tera", "gala",
        "chati", "kamar", "shalwar", "puncha", "note",
        "price", "Date"
    ])
    wb.save(EXCEL_FILE_PATH)


# Now safely load
wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
ws = wb.active


# ========== Routes ==========

@app.route('/')
def login_page():
    return render_template("login.html")

@app.route('/login', methods=["POST"])
def login():
    username = request.form["username"]
    password = request.form["password"]

    if username == USERNAME and password == PASSWORD:
        return redirect(url_for('dashboard'))
    else:
        error = "غلط صارف نام یا پاس ورڈ!"
        return render_template("login.html", error=error)

@app.route('/dashboard')
def dashboard():
    return render_template("dashboard.html")

@app.route('/add-customer')
def add_customer():
    return render_template("add_customer.html")

@app.route('/save_customer', methods=["POST"])
def save_customer():
    data = request.form

    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb.active

    ws.append([

    data.get("id"),
    data.get("name"),
    data.get("phone"),
    data.get("address"),

    "✔" if data.get("collar") else "",
    "✔" if data.get("ban") else "",
    "✔" if data.get("gara_Goal") else "",
    "✔" if data.get("gara_churas") else "",
    "✔" if data.get("fornt_poket") else "",
    "✔" if data.get("site_poket") else "",
    "✔" if data.get("single_salay") else "",
    "✔" if data.get("duble_salay") else "",
    "✔" if data.get("pajama") else "",
    "✔" if data.get("shalwar_checkbox") else "",
    "✔" if data.get("embroidery") else "",
    "✔" if data.get("extra_button") else "",

    data.get("kameez_Length"),
    data.get("bazu"),
    data.get("tera"),
    data.get("gala"),
    data.get("chati"),
    data.get("kamar"),
    data.get("shalwar"),
    data.get("puncha"),
    data.get("note"),
    data.get("price"),
    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    wb.save(EXCEL_FILE_PATH)
    return redirect(url_for("add_customer"))


# ===== view_records route =====
@app.route("/view_records")
def view_records():
    if not os.path.exists(EXCEL_FILE_PATH):
        return "ریکارڈ موجود نہیں!"

    # Excel se data read karo
    df = pd.read_excel(EXCEL_FILE_PATH)
    df = df.loc[:, ~df.columns.duplicated(keep='first')]
    df = df.fillna("خالی")

    # Sirf wo rows jinke paas numeric ID ho
    if "ID" in df.columns:
        df = df[df["ID"].apply(lambda x: str(x).isdigit())]
    else:
        return "ID کالم موجود نہیں!"

    # Column order set karo
    preferred_order = ["ID", "Name", "Phone"]
    all_possible_cols = [
    "Address", "collar", "ban", "gara_Goal", "gara_churas", "fornt_poket",
    "site_poket", "single_salay", "duble_salay", "pajama",
    "shalwar_checkbox",  # ✅ correct
    "Option11", "Option12",
    "kameez_Length", "bazu", "tera", "gala", "chati",
    "kamar", "shalwar",  # ✅ correct
    "puncha", "note", "price", "Date"
    ]   

    remaining_cols = [col for col in all_possible_cols if col in df.columns]
    preferred_order = [col for col in preferred_order if col in df.columns]
    df = df[preferred_order + remaining_cols]

    # Urdu headers mapping
    urdu_headers = {
    "ID": "شناختی نمبر",
    "Name": "نام",
    "Phone": "فون نمبر",
    "Address": "ایڈریس",
    "collar": "کالر",
    "ban": "بین",
    "gara_Goal": "گیرا گول",
    "gara_churas": "گیرا چورس",
    "fornt_poket": "سامنے پاکٹ",
    "site_poket": "سائیڈ پاکٹ",
    "single_salay": "سنگل سلائی",
    "duble_salay": "ڈبل سلائی",
    "pajama": "پاجامہ",
    "shalwar_checkbox": "✔ شلوار",  # ✅ checkbox
    "Option11": "کڑھائی",
    "Option12": "اضافی بٹن",
    "kameez_Length": "قمیض لمبائی",
    "bazu": "بازو",
    "tera": "تیرا",
    "gala": "گلا",
    "chati": "چھاتی",
    "kamar": "کمر",
    "shalwar": "شلوار",  # ✅ input field
    "puncha": "پائنچہ",
    "note": "نوٹ",
    "price": "ادائیگی",
    "Date": "تاریخ"
}



    # Columns list and mappings
    actual_columns = df.columns.tolist()
    headers_urdu = [urdu_headers.get(col, col) for col in actual_columns]
    headers_map = dict(zip(headers_urdu, actual_columns))

    # 👇 Debug: Check what is being sent to HTML
    # print("📋 Columns:", actual_columns)
    # print("📋 Urdu Headers:", headers_urdu)
    # print("📋 Record Sample:", df.head(1).to_dict(orient="records"))

    # Convert to list of records
    records = df.to_dict(orient='records')

    return render_template("view_records.html",
                           headers=headers_urdu,
                           records=records,
                           headers_map=headers_map)


# delete 
@app.route('/delete_customer/<int:customer_id>', methods=["POST"])
def delete_customer(customer_id):
    # Excel file open karo
    df = pd.read_excel(EXCEL_FILE_PATH)

    # Record delete karo jahan شناحتی نمبر match kare
    df = df[df["ID"] != customer_id]

    # Save back to Excel
    df.to_excel(EXCEL_FILE_PATH, index=False)

    flash("ریکارڈ حذف کر دیا گیا!")
    return redirect(url_for("view_records"))



# ========== Run ==========
if __name__ == '__main__':
    app.run(debug=True)




